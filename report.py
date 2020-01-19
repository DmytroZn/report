import pandas as pd
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
from openpyxl import Workbook, load_workbook
from selenium import webdriver

# Завдання 1
# Порахувати суму по колонці **Кількість відпущених упаковок, шт** 
# у файлі **report.xlsx** по виробнику **ПАТ "Фармак"**. 
# Рекомендована бібліотека **pandas**.

class OpenFile():

    def __init__(self, excel_file, sheet_name, num):
        self._excel_file = excel_file
        self._sheet_name = sheet_name
        self._num = num

    def open_f(self):
        opening = pd.read_excel(self._excel_file, sheet_name=self._sheet_name, header=None, skiprows=self._num)
        open_subset = opening[[2,6]]
        open_subset.head()
        dict_sup = open_subset.to_dict()
        return dict_sup

    def open_g(self):
        opening = pd.read_excel(self._excel_file, sheet_name=self._sheet_name, header=None, skiprows=self._num)
        return opening


class OpenDict:
    """ Кількість відпущених упаковок, шт у ПАТ "Фармак" """
    def __init__(self, dict_sup, name):
        self._dict_sup = dict_sup
        self._name = name
  
    def select(self):        
        p = [k for k, v in self._dict_sup[2].items() if v == self._name]
        o = [self._dict_sup[6][i] for i in p]
        return f'Кількість відпущених упаковок, шт у ПАТ "Фармак": {sum(o)}'

    def gett(self):
        return [k for k, v in self._dict_sup[2].items() if v == self._name]
   
u = OpenFile('report.xlsx', 'Report', 19).open_f()
y = OpenDict(u, 'ПАТ "Фармак"').select()
print(y)

# Завдання 2
# Згенерувати новий Excel файл куди ввійдуть зі звіту з першого завдання 
# тільки ті рядки в яких виробник **ТОВ "Кусум Фарм"**. 
# Зробити аркуш **захищеним від редагування**. 
# **Відформатувати** на свій смак. Рекомендована бібліотека **openpyxl**.

class StylesForXl:
    
    def font(self):
        font = Font(name='Calibri',
                            size=10,
                            bold=False,
                            italic=False,
                            vertAlign=None,
                            underline='none',
                            strike=False,
                            color='FF000000')
        return font

    def fill(self):
        fill = PatternFill(
            fill_type='solid',
            start_color='ffffff',
            end_color='ffffff'
            )
        return fill

    def align_center(self):
        align_center=Alignment(
            horizontal='center',
            vertical='bottom',
            text_rotation=0,
            wrapText=True,
            wrap_text=True,
            shrink_to_fit=True,
            indent=1
            )
        return align_center
                   
    def border(self):
        border = Border(
            left=Side(border_style='thin',
                        color='FF000000'),
            right=Side(border_style='thin',
                        color='FF000000'),
            top=Side(border_style='thin',
                        color='FF000000'),
            bottom=Side(border_style='thin',
                        color='FF000000'),
            diagonal=Side(border_style='thin',
                            color='FF000000'),
            diagonal_direction=0,
            outline=Side(border_style='thin',
                            color='FF000000'),
            vertical=Side(border_style='thin',
                            color='FF000000'),
            horizontal=Side(border_style='thin',
                            color='FF000000')
            )
        return border


name_new_file = 'new.xlsx'
name_new_list = 'first_list'

wb = Workbook()
wb.save(name_new_file)
ws = wb.active
ws.title = name_new_list
wb.save(name_new_file)


opening1 = OpenFile("report.xlsx", 'Report', 0).open_g()
opening2 = OpenFile("report.xlsx", 'Report', 19).open_g()
opening3 = OpenFile("report.xlsx", 'Report', 116).open_g()

class Do:
   
    def __init__(self, ws, wb, opening, name_file):
        self._ws = ws
        self._wb = wb
        self._opening = opening
        self._name_file = name_file

    
    def first(self):
        """ num = 0 """
        
        for i in range(19):
            t = self._opening.loc[i].to_list()
            self._ws.append(t)
        return self._wb.save(self._name_file)
    
    def second(self):
        """ num = 19 """
        self._opening.head()
        t = self._opening.loc[0].to_list()
        self._ws.append(t)
        u = opening2
        y = OpenDict(u, 'ТОВ "Кусум Фарм"')
        y = y.gett()
        for i in y:
            t = [self._opening.loc[i].to_list()]
            for row in t:
                self._ws.append(row)
        return self._wb.save(self._name_file)

    def third(self):
        """ num = 116 """
        for i in range(14):
            t = self._opening.loc[i].to_list()
            ws.append(t)
        self._ws['I30'] = '=SUM(I21:I29)'
        self._ws['I35'] = '=SUM(I21:I29)'
        return self._wb.save(self._name_file)


do1 = Do(ws,wb,opening1, name_new_file).first()
do2 = Do(ws, wb, opening2, name_new_file).second()
do3 = Do(ws, wb, opening3, name_new_file).third()
wb.save(name_new_file)

list_of_col = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

for i in list_of_col:
    ws[f'{i}20'].alignment = StylesForXl().align_center()

for i in list_of_col:
    for n in range(1,44):
        ws[f'{i}{n}'].font = StylesForXl().font()

for i in list_of_col:
    for n in range(1,20):
        ws[f'{i}{n}'].fill = StylesForXl().fill()

for i in list_of_col:
    for n in range(30, 44):
        ws[f'{i}{n}'].fill = StylesForXl().fill()
        
for i in list_of_col:
    for n in range(20,30):
        ws[f'{i}{n}'].border = StylesForXl().border()

for i in list_of_col:
    ws.column_dimensions[f'{i}'].width = 15
    ws.row_dimensions[20].height = 75



ws.column_dimensions['A'].width = 5
ws.protection.sheet = True
ws.protection.enable()
wb.save(name_new_file)


# Завдання 3
# З допомогою бібліотеки **Selenium** витягніть з розділу контакти 
# сайту НСЗУ (https://nszu.gov.ua/pro-nszu/kontakti) 
# адресу **українською** та **англійською** мовами

class Pars:

    def __init__(self, url):
        self._url = url

    def go_to_page(self):
        self._driver = webdriver.Chrome("/usr/bin/chromedriver")
        self._driver.get(self._url)
        return self.select_ua()

    def select_ua(self):
        address_ua = self._driver.find_element_by_tag_name('strong')
        print(address_ua.text)
        return self.click_en()

    def click_en(self):
        button_en = self._driver.find_element_by_class_name('lang-item-bottom')
        button_en.click()
        return self.select_en()

    def select_en(self):
        address_en = self._driver.find_element_by_tag_name('strong')
        print(address_en.text)


Pars('https://nszu.gov.ua/pro-nszu/kontakti').go_to_page()


# Завдання 4
# Запропонуйте варіант, як можна реалізувати заміну тексту в pdf файлах. 
# Наприклад як у файлі 
# [Типовий договір](https://nszu.gov.ua/storage/files/19a-35599262-00_1548942590.pdf) 
# замінити **Місто Київ** на **Місто Львів**


# Перше, що спадає на думку це у редакторі замалювати 'Київ' 
# і вставити текст з написом 'Львів';
# Друга ідея це перевести pdf файл у word та зробити зміни 
# і далі зберегти знову у pdf.
# Серудовище для реалізації обох методів 'https://www.sodapdf.com/ru/'

